import os
import sys
import random
import json
import hashlib
import time
import subprocess
import re
import openpyxl
from openpyxl import load_workbook
import tkinter as tk
from tkinter import ttk, messagebox, filedialog, scrolledtext
from PIL import Image, ImageTk
import threading

# 配置信息
ROOT_DIR = "./"  # 题库根目录
PROGRESS_DIR = "progress"  # 进度保存目录

def install_package(package):
    """安装必要的Python包"""
    try:
        __import__(package)
    except ImportError:
        print(f"正在安装 {package}...")
        subprocess.check_call([sys.executable, "-m", "pip", "install", package])
        print(f"{package} 安装完成!")

def scan_subjects():
    """扫描题库目录，返回包含xlsx文件的科目列表"""
    subjects = []
    for entry in os.listdir(ROOT_DIR):
        full_path = os.path.join(ROOT_DIR, entry)
        if os.path.isdir(full_path):
            # 检查目录是否包含xlsx文件
            has_xlsx = False
            for root, dirs, files in os.walk(full_path):
                for file in files:
                    if file.endswith(".xlsx") and not file.startswith("~$"):
                        has_xlsx = True
                        break
                if has_xlsx:
                    break
            if has_xlsx:
                subjects.append(entry)
    return subjects

def scan_question_files(subject):
    """扫描指定科目下的题库文件"""
    subject_dir = os.path.join(ROOT_DIR, subject)
    question_files = []

    for root, dirs, files in os.walk(subject_dir):
        for file in files:
            if file.endswith(".xlsx") and not file.startswith("~$"):
                full_path = os.path.join(root, file)
                question_files.append(full_path)

    return question_files

def parse_question_file(file_path):
    """解析题库文件，返回题目列表"""
    wb = load_workbook(file_path)
    sheet = wb.active
    questions = []

    headers = [cell.value for cell in sheet[1]]

    image_columns = ["附图", "图片", "image", "Image", "picture", "Picture"]
    image_col = None
    for col in image_columns:
        if col in headers:
            image_col = col
            break

    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not any(row):  # 跳过空行
            continue

        question = {}
        for i, value in enumerate(row):
            if i >= len(headers):
                break
            header = headers[i]
            if header and value is not None:
                question[header] = value

        # 处理选择题选项 - 修复逻辑
        options = []
        raw_options = []

        # 获取选项列的值
        options_value = question.get("选项", "")

        # 处理多选题答案格式
        if question.get("题型") == "多选题":
            # 答案格式为 "A | B | C"
            answer_value = question.get("答案", "")
            if isinstance(answer_value, str) and "|" in answer_value:
                question["answer_parts"] = [part.strip() for part in answer_value.split("|")]
            else:
                question["answer_parts"] = [answer_value.strip()]

        if options_value:
            # 情况1：选项是列表格式的字符串
            if isinstance(options_value, str) and options_value.startswith('[') and options_value.endswith(']'):
                try:
                    # 尝试解析为Python列表
                    parsed_options = eval(options_value)
                    if isinstance(parsed_options, list):
                        raw_options = parsed_options
                except:
                    # 解析失败，按竖线分割处理
                    raw_options = [opt.strip() for opt in options_value.strip("[]").split("|")]

            # 情况2：选项是用竖线分隔的字符串
            elif isinstance(options_value, str) and "|" in options_value:
                raw_options = [opt.strip() for opt in options_value.split("|")]

            # 情况3：选项是单个字符串
            elif isinstance(options_value, str):
                raw_options = [options_value.strip()]

            # 清理每个选项格式
            for opt in raw_options:
                # 清理选项格式：移除开头的字母和标点
                clean_opt = re.sub(r"^[A-Za-z][\.\s]*", "", opt).strip()
                options.append(clean_opt)

        if image_col and image_col in question:
            image_path = question[image_col]
            if image_path and isinstance(image_path, str) and image_path.strip():
                # 处理相对路径（相对于Excel文件所在目录）
                base_dir = os.path.dirname(file_path)
                abs_path = os.path.join(base_dir, image_path.strip())
                question["image_path"] = abs_path

        # 判断题特殊处理
        elif question.get("题型") == "判断题":
            raw_options = ["正确", "错误"]
            options = ["正确", "错误"]

        question["options"] = options
        question["raw_options"] = raw_options

        questions.append(question)

    return questions

def normalize_answer(answer):
    """标准化答案格式"""
    if isinstance(answer, str):
        answer = answer.strip()
        # 处理判断题
        if "正确" in answer or "对" in answer or "是" in answer or "T" in answer or "t" in answer:
            return "正确"
        elif "错误" in answer or "错" in answer or "否" in answer or "F" in answer or "f" in answer:
            return "错误"
    return answer

def get_progress_file_path(question_file):
    """获取进度文件路径"""
    if not os.path.exists(PROGRESS_DIR):
        os.makedirs(PROGRESS_DIR)

    file_hash = hashlib.md5(question_file.encode()).hexdigest()
    return os.path.join(PROGRESS_DIR, f"{file_hash}.json")

def load_progress(question_file):
    """加载进度信息"""
    progress_file = get_progress_file_path(question_file)
    if os.path.exists(progress_file):
        try:
            with open(progress_file, "r", encoding="utf-8") as f:
                return json.load(f)
        except:
            pass

    # 默认进度信息
    return {
        "total_questions": 0,
        "answered": {},
        "wrong_questions": [],
        "current_index": 0,
        "correct_count": 0,
        "wrong_count": 0
    }

def save_progress(question_file, progress):
    """保存进度信息"""
    progress_file = get_progress_file_path(question_file)
    with open(progress_file, "w", encoding="utf-8") as f:
        json.dump(progress, f, ensure_ascii=False, indent=2)

def check_answer(question, user_answer):
    """检查答案是否正确"""
    correct_answer = normalize_answer(question.get("答案", ""))
    user_answer = normalize_answer(user_answer)

    # 选择题检查
    if (question.get("题型") == "选择题" or question.get("题型") == "判断题") and question.get("options"):
        letters = ["A", "B", "C", "D", "E", "F", "G", "H"]
        options = question.get("raw_options", [])

        # 查找用户选择的选项文本
        user_option_text = None
        for i, letter in enumerate(letters):
            if i < len(options) and user_answer.upper() == letter:
                user_option_text = options[i]
                break

        # 查找标准答案对应的选项文本
        correct_option_text = None
        for i, letter in enumerate(letters):
            if i < len(options) and correct_answer.upper() == letter:
                correct_option_text = options[i]
                break

        # 如果找不到则使用原始答案
        if user_option_text is None:
            user_option_text = user_answer
        if correct_option_text is None:
            correct_option_text = correct_answer
        # 使用转换后的文本进行比较
        return user_option_text.strip() == correct_option_text.strip(), correct_option_text.strip()

    # 多选题检查
    if question.get("题型") == "多选题":
        # 用户答案格式为 "A | B | C"
        user_parts = [part.strip().upper() for part in user_answer.split("|")] if "|" in user_answer else [user_answer.strip().upper()]
        correct_parts = [part.strip().upper() for part in question.get("answer_parts", [])]

        # 比较答案（顺序无关）
        if set(user_parts) == set(correct_parts):
            return True, " | ".join(correct_parts)
        return False, " | ".join(correct_parts)

    # 填空题处理（多个空）
    if question.get("题型") == "填空题":
        # 使用特定分隔符 || 分割答案
        if " | " in correct_answer:
            correct_parts = [part.strip() for part in correct_answer.split("||")]
        else:
            correct_parts = [correct_answer.strip()]

        if " | " in user_answer:
            user_parts = [part.strip() for part in user_answer.split("||")]
        else:
            user_parts = [user_answer.strip()]

        if len(correct_parts) != len(user_parts):
            return False, correct_answer

        for c, u in zip(correct_parts, user_parts):
            if c != u:
                return False, correct_answer

        return True, correct_answer

    # 其他题型直接比较
    return user_answer == correct_answer, correct_answer

class ExamApp:
    def __init__(self, root):
        self.root = root
        self.root.title("智能刷题系统")
        self.root.geometry("900x700")
        self.root.configure(bg="#f0f0f0")

        # 初始化变量
        self.subjects = []
        self.question_files = []
        self.questions = []
        self.progress = {}
        self.question_order = []
        self.current_index = 0
        self.selected_subject = ""
        self.selected_file = ""
        self.default_wait_seconds = 5  # 默认等待时间（秒）
        self.showing_answer = False     # 是否正在显示答案
        self.review_mode = False        # 背题模式标志
        self.filter_types = ["全部"]    # 题型筛选选项
        self.selected_filter = "全部"   # 当前选中的题型筛选
        self.filter_menu_open = False   # 筛选菜单是否打开

        # 创建主框架
        self.create_welcome_frame()

        # 确保安装所需包
        threading.Thread(target=self.install_required_packages, daemon=True).start()

        self.result_label = None      # 结果标签
        self.countdown_label = None   # 倒计时标签
        self.countdown_id = None      # 倒计时任务ID
        self.countdown_seconds = 5    # 倒计时秒数
        self.wait_time_label = None   # 等待时间显示标签

        self.multi_select_vars = {}  # 存储多选题选项状态
        self.multi_select_frame = None  # 多选题选项框架

    def install_required_packages(self):
        install_package("openpyxl")
        install_package("pillow")

    def create_welcome_frame(self):
        """创建欢迎界面"""
        self.clear_frame()

        # 标题
        title_label = tk.Label(self.root, text="智能刷题系统", font=("微软雅黑", 24, "bold"), bg="#f0f0f0", fg="#333")
        title_label.pack(pady=20)

        # 图标
        try:
            icon_img = Image.open("icon.png") if os.path.exists("icon.png") else None
            if icon_img:
                icon_img = icon_img.resize((150, 150), Image.LANCZOS)
                self.icon_photo = ImageTk.PhotoImage(icon_img)
                icon_label = tk.Label(self.root, image=self.icon_photo, bg="#f0f0f0")
                icon_label.pack(pady=10)
        except:
            pass

        # 说明文本
        desc_label = tk.Label(self.root, text="选择科目开始刷题练习", font=("微软雅黑", 14), bg="#f0f0f0", fg="#555")
        desc_label.pack(pady=10)

        # 科目选择按钮
        subject_btn = tk.Button(self.root, text="选择科目", command=self.show_subject_selection,
                                font=("微软雅黑", 12), bg="#4CAF50", fg="white", padx=20, pady=10)
        subject_btn.pack(pady=20)

        # 进度管理按钮
        progress_btn = tk.Button(self.root, text="进度管理", command=self.show_progress_management,
                                 font=("微软雅黑", 12), bg="#2196F3", fg="white", padx=20, pady=10)
        progress_btn.pack(pady=10)

        # 退出按钮
        exit_btn = tk.Button(self.root, text="退出系统", command=self.root.quit,
                             font=("微软雅黑", 12), bg="#F44336", fg="white", padx=20, pady=10)
        exit_btn.pack(pady=10)

    def show_subject_selection(self):
        """显示科目选择界面"""
        self.clear_frame()

        # 标题
        title_label = tk.Label(self.root, text="选择科目", font=("微软雅黑", 20, "bold"), bg="#f0f0f0")
        title_label.pack(pady=20)

        # 扫描科目
        self.subjects = scan_subjects()

        if not self.subjects:
            no_subject_label = tk.Label(self.root, text="未找到任何包含题目的科目!", font=("微软雅黑", 12), bg="#f0f0f0", fg="red")
            no_subject_label.pack(pady=20)

            back_btn = tk.Button(self.root, text="返回", command=self.create_welcome_frame,
                                 font=("微软雅黑", 12), bg="#2196F3", fg="white")
            back_btn.pack(pady=10)
            return

        # 科目选择框架
        subject_frame = tk.Frame(self.root, bg="#f0f0f0")
        subject_frame.pack(pady=10, fill=tk.BOTH, expand=True)

        # 科目列表
        for i, subject in enumerate(self.subjects):
            btn = tk.Button(subject_frame, text=subject, command=lambda s=subject: self.select_subject(s),
                            font=("微软雅黑", 12), bg="#E0E0E0", width=30, height=2)
            btn.grid(row=i, column=0, padx=20, pady=10, sticky="ew")

        # 返回按钮
        back_btn = tk.Button(self.root, text="返回", command=self.create_welcome_frame,
                             font=("微软雅黑", 12), bg="#2196F3", fg="white")
        back_btn.pack(pady=10)

    def select_subject(self, subject):
        """选择科目并显示题库文件"""
        self.selected_subject = subject
        self.question_files = scan_question_files(subject)

        if not self.question_files:
            messagebox.showerror("错误", f"在 '{subject}' 中未找到题库文件!")
            return

        self.clear_frame()

        # 标题
        title_label = tk.Label(self.root, text=f"选择题库文件 - {subject}", font=("微软雅黑", 20, "bold"), bg="#f0f0f0")
        title_label.pack(pady=20)

        # 文件列表框架
        file_frame = tk.Frame(self.root, bg="#f0f0f0")
        file_frame.pack(pady=10, fill=tk.BOTH, expand=True)

        # 文件列表
        for i, file_path in enumerate(self.question_files):
            file_name = os.path.basename(file_path)
            btn = tk.Button(file_frame, text=file_name, command=lambda f=file_path: self.select_file(f),
                            font=("微软雅黑", 11), bg="#E0E0E0", width=40, height=1, anchor="w")
            btn.grid(row=i, column=0, padx=20, pady=5, sticky="w")

        # 返回按钮
        back_btn = tk.Button(self.root, text="返回", command=self.show_subject_selection,
                             font=("微软雅黑", 12), bg="#2196F3", fg="white")
        back_btn.pack(pady=10)

    def select_file(self, file_path):
        """选择题库文件并开始答题"""
        self.selected_file = file_path

        try:
            self.questions = parse_question_file(file_path)
            if not self.questions:
                messagebox.showerror("错误", "题库中没有题目!")
                return
        except Exception as e:
            messagebox.showerror("加载失败", f"加载题库失败: {e}")
            return

        # 加载进度
        self.progress = load_progress(file_path)
        self.progress["total_questions"] = len(self.questions)

        # 提取所有题型
        all_types = set()
        for q in self.questions:
            q_type = q.get("题型", "未知题型")
            if q_type:
                all_types.add(q_type)

        # 更新题型筛选选项
        self.filter_types = ["全部"] + sorted(list(all_types))
        self.selected_filter = "全部"

        # 准备题目顺序
        self.generate_question_order()

        self.current_index = 0
        self.show_question()

    def generate_question_order(self):
        """生成题目顺序（考虑题型筛选）"""
        # 优先错题
        self.question_order = self.progress["wrong_questions"][:]

        # 添加未做过的题目
        all_indices = set(range(len(self.questions)))
        answered_indices = set(int(k) for k in self.progress["answered"].keys())
        unanswered_indices = list(all_indices - answered_indices)
        random.shuffle(unanswered_indices)

        self.question_order.extend(unanswered_indices)

        # 如果没有错题或未做题，使用所有题目
        if not self.question_order:
            self.question_order = list(range(len(self.questions)))
            random.shuffle(self.question_order)

        # 应用题型筛选
        if self.selected_filter != "全部":
            filtered_order = []
            for idx in self.question_order:
                q_type = self.questions[idx].get("题型", "未知题型")
                if q_type == self.selected_filter:
                    filtered_order.append(idx)
            self.question_order = filtered_order

    def show_question(self):
        """显示当前题目"""
        if self.countdown_id:
            self.root.after_cancel(self.countdown_id)
            self.countdown_id = None

        self.clear_frame()
        self.showing_answer = False  # 重置答案显示状态

        if self.current_index >= len(self.question_order):
            self.show_results()
            return

        q_index = self.question_order[self.current_index]
        self.current_question = self.questions[q_index]

        # 主框架
        main_frame = tk.Frame(self.root, bg="#f0f0f0")
        main_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=10)

        # 题目信息
        info_frame = tk.Frame(main_frame, bg="#f0f0f0")
        info_frame.pack(fill=tk.X, pady=5)

        # 题型显示和筛选按钮
        type_frame = tk.Frame(info_frame, bg="#f0f0f0")
        type_frame.pack(side=tk.RIGHT, padx=10)

        # 题型筛选菜单
        filter_btn = tk.Menubutton(type_frame, text="题型筛选 ▼",
                                   font=("微软雅黑", 10), bg="#E0E0E0",
                                   relief=tk.RAISED, padx=5, pady=2)
        filter_btn.menu = tk.Menu(filter_btn, tearoff=0)
        filter_btn["menu"] = filter_btn.menu

        # 添加筛选选项
        for t in self.filter_types:
            filter_btn.menu.add_radiobutton(
                label=t,
                command=lambda t=t: self.apply_type_filter(t),
                variable=tk.StringVar(value=self.selected_filter)
            )

        filter_btn.pack(side=tk.RIGHT, padx=5)

        # 当前题型显示
        q_type = self.current_question.get("题型", "未知题型")
        type_label = tk.Label(type_frame, text=f"题型: {q_type}",
                              font=("微软雅黑", 12), bg="#f0f0f0")
        type_label.pack(side=tk.RIGHT, padx=5)

        # 背题模式切换按钮
        review_btn = tk.Button(type_frame, text="背题模式" if not self.review_mode else "练习模式",
                               font=("微软雅黑", 10),
                               bg="#9C27B0" if self.review_mode else "#E0E0E0",
                               fg="white" if self.review_mode else "black",
                               command=self.toggle_review_mode)
        review_btn.pack(side=tk.RIGHT, padx=5)

        # 题目编号
        tk.Label(info_frame, text=f"题目 {self.current_index+1}/{len(self.question_order)}",
                 font=("微软雅黑", 12), bg="#f0f0f0").pack(side=tk.LEFT)

        # 问题内容
        question_frame = tk.LabelFrame(main_frame, text="问题", font=("微软雅黑", 12, "bold"),
                                       bg="#f0f0f0", padx=10, pady=10)
        question_frame.pack(fill=tk.BOTH, expand=True, pady=10)

        question_text = scrolledtext.ScrolledText(question_frame, font=("微软雅黑", 12),
                                                  wrap=tk.WORD, height=6)
        question_text.insert(tk.INSERT, self.current_question.get('问题', ''))
        question_text.config(state=tk.DISABLED)
        question_text.pack(fill=tk.BOTH, expand=True)

        # 背题模式下直接显示答案
        if self.review_mode:

            if self.current_question.get("题型") in ["选择题", "判断题", "多选题"] and self.current_question.get("options"):
                options_frame = tk.LabelFrame(main_frame, text="选项与答案",
                                              font=("微软雅黑", 12, "bold"),
                                              bg="#f0f0f0", padx=10, pady=10)
                options_frame.pack(fill=tk.BOTH, expand=True, pady=10)

                # 获取正确答案
                correct_answer = normalize_answer(self.current_question.get("答案", ""))
                q_type = self.current_question.get("题型", "")
                raw_options = self.current_question.get("raw_options", [])
                letters = ["A", "B", "C", "D", "E", "F", "G", "H"]

                # 确定正确选项（支持多选题）
                correct_answers = []
                if q_type == "多选题":
                    # 处理多选题答案格式（可能包含多个选项）
                    if "|" in correct_answer:
                        correct_parts = [part.strip() for part in correct_answer.split("|")]
                    else:
                        correct_parts = [correct_answer.strip()]

                    # 将答案转换为选项字母
                    for part in correct_parts:
                        if part in letters:  # 如果答案已经是字母
                            correct_answers.append(part)
                        else:  # 如果答案是文本，查找对应的选项
                            for idx, opt in enumerate(raw_options):
                                if opt.strip() == part.strip() and idx < len(letters):
                                    correct_answers.append(letters[idx])
                else:
                    # 单选题和判断题处理
                    if correct_answer in letters:  # 如果答案已经是字母
                        correct_answers = [correct_answer]
                    else:  # 如果答案是文本，查找对应的选项
                        for idx, opt in enumerate(raw_options):
                            if opt.strip() == correct_answer.strip() and idx < len(letters):
                                correct_answers = [letters[idx]]

                # 显示所有选项，正确选项用绿色标记
                for i, option_text in enumerate(raw_options):
                    if i >= len(letters):
                        break

                    letter = letters[i]
                    # 创建选项框架
                    option_frame = tk.Frame(options_frame, bg="#f0f0f0")
                    option_frame.pack(fill=tk.X, pady=2)

                    # 选项标签
                    label_text = f"{letter}. {option_text}"
                    label = tk.Label(option_frame, text=label_text,
                                     font=("微软雅黑", 11),
                                     bg="#f0f0f0",
                                     anchor="w")

                    # 如果是正确选项，使用绿色标记
                    if letter in correct_answers:
                        label.config(fg="green", font=("微软雅黑", 11, "bold"))

                    label.pack(side=tk.LEFT, fill=tk.X, expand=True)

            else:
                answer_frame = tk.LabelFrame(main_frame, text="标准答案",
                                             font=("微软雅黑", 12, "bold"),
                                             bg="#f0f0f0", padx=10, pady=10)
                answer_frame.pack(fill=tk.BOTH, expand=True, pady=10)

                answer_text = scrolledtext.ScrolledText(answer_frame, font=("微软雅黑", 12),
                                                        wrap=tk.WORD, height=4)
                # correct_answer = normalize_answer(self.current_question.get("答案", ""))
                # print()
                # print(" ")
                # print(correct_answer)
                answer_text.insert(tk.INSERT, self.current_question.get("答案","???"))
                answer_text.config(state=tk.DISABLED)
                answer_text.pack(fill=tk.BOTH, expand=True)

        '''# 背题模式下直接显示答案
        if self.review_mode:
            answer_frame = tk.LabelFrame(main_frame, text="标准答案",
                                         font=("微软雅黑", 12, "bold"),
                                         bg="#f0f0f0", padx=10, pady=10)
            answer_frame.pack(fill=tk.BOTH, expand=True, pady=10)

            answer_text = scrolledtext.ScrolledText(answer_frame, font=("微软雅黑", 12),
                                                    wrap=tk.WORD, height=4)
            correct_answer = normalize_answer(self.current_question.get("答案", ""))

            # 选择题处理：显示选项字母
            if self.current_question.get("题型") in ["选择题", "判断题", "多选题"] and self.current_question.get("options"):
                letters = ["A", "B", "C", "D", "E", "F", "G", "H"]
                options = self.current_question.get("raw_options", [])

                for i, opt in enumerate(options):
                    if opt.strip() == correct_answer.strip() and i < len(letters):
                        correct_answer = f"{letters[i]}. {opt}"
                        break

            answer_text.insert(tk.INSERT, correct_answer)
            answer_text.config(state=tk.DISABLED)
            answer_text.pack(fill=tk.BOTH, expand=True)'''

        # 显示图片
        if "image_path" in self.current_question and self.current_question["image_path"]:
            try:
                image_path = self.current_question["image_path"]

                # 加载图片并调整大小
                img = Image.open(image_path)
                width, height = img.size
                max_width = 600
                if width > max_width:
                    ratio = max_width / width
                    new_height = int(height * ratio)
                    img = img.resize((max_width, new_height), Image.LANCZOS)

                self.photo = ImageTk.PhotoImage(img)

                # 创建图片显示区域
                image_frame = tk.LabelFrame(main_frame, text="附图",
                                            font=("微软雅黑", 12, "bold"), bg="#f0f0f0")
                image_frame.pack(fill=tk.BOTH, expand=True, pady=10)

                img_label = tk.Label(image_frame, image=self.photo, bg="#f0f0f0")
                img_label.pack(padx=10, pady=10)

                # 添加图片路径提示
                path_label = tk.Label(image_frame, text=f"图片路径: {image_path}",
                                      font=("微软雅黑", 9), fg="#666", bg="#f0f0f0")
                path_label.pack(side=tk.BOTTOM, padx=10, pady=5)

            except Exception as e:
                error_frame = tk.LabelFrame(main_frame, text="图片加载失败",
                                            font=("微软雅黑", 12, "bold"), bg="#f0f0f0", fg="red")
                error_frame.pack(fill=tk.X, pady=10)

                error_label = tk.Label(error_frame, text=f"无法加载图片: {str(e)}",
                                       font=("微软雅黑", 10), fg="red", bg="#f0f0f0")
                error_label.pack(padx=10, pady=5)

        # 非背题模式显示答题区域
        if not self.review_mode:
            # 选项（选择题）
            options = self.current_question.get("options", [])
            if options and self.current_question.get("题型") != "多选题":
                # 非多选题使用按钮
                options_frame = tk.LabelFrame(main_frame, text="选项", font=("微软雅黑", 12, "bold"),
                                              bg="#f0f0f0", padx=10, pady=10)
                options_frame.pack(fill=tk.BOTH, expand=True, pady=10)

                letters = ["A", "B", "C", "D", "E", "F", "G", "H"]
                for i, opt in enumerate(options):
                    if i < len(letters):
                        btn = tk.Button(options_frame, text=f"{letters[i]}. {opt}",
                                        command=lambda l=letters[i]: self.check_answer_wrapper(l),
                                        font=("微软雅黑", 11), bg="#E0E0E0", width=40, anchor="w")
                        btn.pack(pady=5, padx=10, anchor="w")

            elif options and self.current_question.get("题型") == "多选题":
                # 多选题使用复选框
                self.multi_select_vars = {}  # 重置选项状态

                options_frame = tk.LabelFrame(main_frame, text="选项（可多选）",
                                              font=("微软雅黑", 12, "bold"),
                                              bg="#f0f0f0", padx=10, pady=10)
                options_frame.pack(fill=tk.BOTH, expand=True, pady=10)
                self.multi_select_frame = options_frame  # 保存引用

                letters = ["A", "B", "C", "D", "E", "F", "G", "H"]
                for i, opt in enumerate(options):
                    if i < len(letters):
                        var = tk.BooleanVar()
                        self.multi_select_vars[letters[i]] = var

                        cb = tk.Checkbutton(options_frame, text=f"{letters[i]}. {opt}",
                                            variable=var, font=("微软雅黑", 11),
                                            bg="#f0f0f0", anchor="w")
                        cb.pack(fill=tk.X, pady=3, padx=10)

                # 添加多选题提交按钮
                multi_submit_frame = tk.Frame(main_frame, bg="#f0f0f0")
                multi_submit_frame.pack(pady=5)

                submit_btn = tk.Button(multi_submit_frame, text="提交多选题答案",
                                       command=self.submit_multi_choice,
                                       font=("微软雅黑", 12), bg="#4CAF50", fg="white")
                submit_btn.pack(pady=5)

            if self.current_question.get("题型") in ["解答题", "简答题"]:
                answer_frame = tk.LabelFrame(main_frame, text="您的解答",
                                             font=("微软雅黑", 12, "bold"), bg="#f0f0f0")
                answer_frame.pack(fill=tk.BOTH, expand=True, pady=10)

                self.answer_text = scrolledtext.ScrolledText(answer_frame, font=("微软雅黑", 12),
                                                             wrap=tk.WORD, height=8)
                self.answer_text.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)

                # 添加手动评分按钮
                self.manual_check_frame = tk.Frame(main_frame, bg="#f0f0f0")
                self.manual_check_frame.pack(pady=10)

                tk.Button(self.manual_check_frame, text="提交并标记正确",
                          command=lambda: self.manual_check_answer(True),
                          font=("微软雅黑", 12), bg="#4CAF50", fg="white").pack(side=tk.LEFT, padx=5)

                tk.Button(self.manual_check_frame, text="提交并标记错误",
                          command=lambda: self.manual_check_answer(False), font=("微软雅黑", 12), bg="#F44336", fg="white").pack(side=tk.LEFT, padx=5)
            else:
                # 答案输入（非选择题）
                if not options:
                    answer_frame = tk.Frame(main_frame, bg="#f0f0f0")
                    answer_frame.pack(fill=tk.X, pady=10)

                    tk.Label(answer_frame, text="答案:", font=("微软雅黑", 12), bg="#f0f0f0").pack(side=tk.LEFT)

                    self.answer_entry = tk.Entry(answer_frame, font=("微软雅黑", 12), width=30)
                    self.answer_entry.pack(side=tk.LEFT, padx=10)
                    self.answer_entry.bind("<Return>", lambda event: self.check_answer_wrapper(self.answer_entry.get()))

                    submit_btn = tk.Button(answer_frame, text="提交",
                                           command=lambda: self.check_answer_wrapper(self.answer_entry.get()),
                                           font=("微软雅黑", 12), bg="#4CAF50", fg="white")
                    submit_btn.pack(side=tk.LEFT)

        # 导航按钮
        nav_frame = tk.Frame(main_frame, bg="#f0f0f0")
        nav_frame.pack(fill=tk.X, pady=10)

        if self.current_index > 0:
            prev_btn = tk.Button(nav_frame, text="上一题", command=self.prev_question,
                                 font=("微软雅黑", 12), bg="#2196F3", fg="white")
            prev_btn.pack(side=tk.LEFT, padx=10)

        skip_btn = tk.Button(nav_frame, text="跳过", command=self.next_question,
                             font=("微软雅黑", 12), bg="#FF9800", fg="white")
        skip_btn.pack(side=tk.LEFT, padx=10)

        exit_btn = tk.Button(nav_frame, text="退出", command=self.create_welcome_frame,
                             font=("微软雅黑", 12), bg="#F44336", fg="white")
        exit_btn.pack(side=tk.RIGHT, padx=10)

        # 非背题模式显示查看答案按钮
        if not self.review_mode:
            view_answer_btn = tk.Button(nav_frame, text="查看答案", command=self.show_answer,
                                        font=("微软雅黑", 12), bg="#9C27B0", fg="white")
            view_answer_btn.pack(side=tk.LEFT, padx=10)

            # 速度控制按钮
            speed_frame = tk.Frame(nav_frame, bg="#f0f0f0")
            speed_frame.pack(side=tk.RIGHT, padx=10)

            # 加速按钮
            speed_up_btn = tk.Button(speed_frame, text="加速", command=self.speed_up,
                                     font=("微软雅黑", 10), bg="#FF5722", fg="white")
            speed_up_btn.pack(side=tk.LEFT, padx=(0, 5))

            # 等待时间显示
            self.wait_time_label = tk.Label(speed_frame, text=f"等待: {self.default_wait_seconds}秒",
                                            font=("微软雅黑", 10), bg="#f0f0f0")
            self.wait_time_label.pack(side=tk.LEFT, padx=5)

            # 减速按钮
            speed_down_btn = tk.Button(speed_frame, text="减速", command=self.speed_down,
                                       font=("微软雅黑", 10), bg="#3F51B5", fg="white")
            speed_down_btn.pack(side=tk.LEFT, padx=(5, 0))

        result_frame = tk.Frame(main_frame, bg="#f0f0f0")
        result_frame.pack(fill=tk.X, pady=10)

        # 结果标签 - 初始为空
        self.result_label = tk.Label(result_frame, text="", font=("微软雅黑", 14), bg="#f0f0f0")
        self.result_label.pack(side=tk.LEFT)

        # 倒计时标签 - 初始为空
        self.countdown_label = tk.Label(result_frame, text="", font=("微软雅黑", 12), fg="#666", bg="#f0f0f0")
        self.countdown_label.pack(side=tk.RIGHT)

    def toggle_review_mode(self):
        """切换背题模式"""
        self.review_mode = not self.review_mode
        self.show_question()

    def apply_type_filter(self, filter_type):
        """应用题型筛选"""
        self.selected_filter = filter_type
        self.current_index = 0
        self.generate_question_order()
        self.show_question()

    def submit_multi_choice(self):
        """提交多选题答案"""
        selected_letters = []
        for letter, var in self.multi_select_vars.items():
            if var.get():
                selected_letters.append(letter)

        if not selected_letters:
            messagebox.showwarning("提示", "请至少选择一个选项")
            return

        # 将选择的选项按字母排序并用" | "连接
        selected_letters.sort()
        user_answer = " | ".join(selected_letters)

        # 检查答案
        self.check_answer_wrapper(user_answer)

        # 清除选项状态
        for var in self.multi_select_vars.values():
            var.set(False)

    def manual_check_answer(self, is_correct):
        """手动评分处理"""
        q_index = self.question_order[self.current_index]
        user_answer = self.answer_text.get("1.0", tk.END).strip()

        # 更新进度
        self.progress["answered"][str(q_index)] = {
            "user_answer": user_answer,
            "is_correct": is_correct,
            "timestamp": time.time()
        }

        if is_correct:
            self.progress["correct_count"] = self.progress.get("correct_count", 0) + 1
            if q_index in self.progress["wrong_questions"]:
                self.progress["wrong_questions"].remove(q_index)
        else:
            self.progress["wrong_count"] = self.progress.get("wrong_count", 0) + 1
            if q_index not in self.progress["wrong_questions"]:
                self.progress["wrong_questions"].append(q_index)

        save_progress(self.selected_file, self.progress)
        self.next_question()

    def show_answer(self):
        """显示当前题目的正确答案"""
        if self.showing_answer:
            return

        self.showing_answer = True
        correct_answer = normalize_answer(self.current_question.get("答案", ""))

        # 选择题处理：显示选项字母
        if self.current_question.get("题型") in ["选择题", "判断题"] and self.current_question.get("options"):
            letters = ["A", "B", "C", "D", "E", "F", "G", "H"]
            options = self.current_question.get("raw_options", [])

            for i, opt in enumerate(options):
                if opt.strip() == correct_answer.strip() and i < len(letters):
                    correct_answer = f"{letters[i]}. {opt}"
                    break

        self.result_label.config(text=f"正确答案: {correct_answer}", fg="blue")

        # 如果正在倒计时，取消倒计时
        if self.countdown_id:
            self.root.after_cancel(self.countdown_id)
            self.countdown_id = None
            self.countdown_label.config(text="")

    def speed_up(self):
        """减少等待时间"""
        if self.default_wait_seconds > 0:
            self.default_wait_seconds -= 1
            self.update_wait_label()

    def speed_down(self):
        """增加等待时间"""
        if self.default_wait_seconds < 60:
            self.default_wait_seconds += 1
            self.update_wait_label()

    def update_wait_label(self):
        """更新等待时间标签"""
        if hasattr(self, 'wait_time_label') and self.wait_time_label:
            self.wait_time_label.config(text=f"等待: {self.default_wait_seconds}秒")

    def check_answer_wrapper(self, answer):
        """检查答案并显示结果"""
        is_correct, correct_answer = check_answer(self.current_question, answer)
        q_index = self.question_order[self.current_index]

        # 更新进度
        self.progress["answered"][str(q_index)] = {
            "user_answer": answer,
            "is_correct": is_correct,
            "timestamp": time.time()
        }

        if is_correct:
            self.progress["correct_count"] = self.progress.get("correct_count", 0) + 1
            # 从错题列表中移除
            if q_index in self.progress["wrong_questions"]:
                self.progress["wrong_questions"].remove(q_index)
            message = "✓ 回答正确!"
            color = "green"
        else:
            self.progress["wrong_count"] = self.progress.get("wrong_count", 0) + 1
            # 添加到错题列表
            if q_index not in self.progress["wrong_questions"]:
                self.progress["wrong_questions"].append(q_index)
            message = f"✗ 回答错误!\n正确答案: {correct_answer}"
            color = "red"

        save_progress(self.selected_file, self.progress)

        # 在界面内显示结果
        if is_correct:
            self.result_label.config(text="✓ 回答正确!", fg="green")
        else:
            self.result_label.config(text=f"✗ 回答错误! 正确答案: {correct_answer}", fg="red")

        # 无论对错都启动倒计时（但只有正确答题会跳转）
        self.start_countdown(is_correct)

    def start_countdown(self, is_correct):
        """启动倒计时"""
        # 清除之前的倒计时
        if self.countdown_id:
            self.root.after_cancel(self.countdown_id)

        # 只有答对才需要倒计时跳转
        if not is_correct:
            self.countdown_label.config(text="")
            return

        # 更新倒计时显示
        self.countdown_seconds = self.default_wait_seconds
        self.update_countdown()

    def update_countdown(self):
        """更新倒计时显示并触发跳转"""
        if self.countdown_seconds > 0:
            self.countdown_label.config(text=f"{self.countdown_seconds}秒后自动跳转")
            self.countdown_seconds -= 1
            self.countdown_id = self.root.after(1000, self.update_countdown)
        else:
            self.countdown_label.config(text="正在跳转...")
            self.next_question()

    def prev_question(self):
        """显示上一题"""
        if self.countdown_id:
            self.root.after_cancel(self.countdown_id)
            self.countdown_id = None

        if self.current_index > 0:
            self.current_index -= 1
            self.show_question()

    def next_question(self):
        """显示下一题"""
        if self.countdown_id:
            self.root.after_cancel(self.countdown_id)
            self.countdown_id = None

        self.current_index += 1
        self.show_question()

    def show_results(self):
        """显示答题结果统计"""
        self.clear_frame()

        # 主框架
        main_frame = tk.Frame(self.root, bg="#f0f0f0")
        main_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)

        # 标题
        tk.Label(main_frame, text="答题完成", font=("微软雅黑", 20, "bold"), bg="#f0f0f0").pack(pady=20)

        # 统计信息
        total_answered = len(self.progress["answered"])
        correct_count = self.progress.get("correct_count", 0)
        wrong_count = self.progress.get("wrong_count", 0)
        accuracy = correct_count / total_answered * 100 if total_answered > 0 else 0

        stats_frame = tk.Frame(main_frame, bg="#f0f0f0")
        stats_frame.pack(pady=20)

        tk.Label(stats_frame, text=f"总答题数: {total_answered}", font=("微软雅黑", 14), bg="#f0f0f0").grid(row=0, column=0, sticky="w", pady=5)
        tk.Label(stats_frame, text=f"正确: {correct_count}", font=("微软雅黑", 14), fg="green", bg="#f0f0f0").grid(row=1, column=0, sticky="w", pady=5)
        tk.Label(stats_frame, text=f"错误: {wrong_count}", font=("微软雅黑", 14), fg="red", bg="#f0f0f0").grid(row=2, column=0, sticky="w", pady=5)
        tk.Label(stats_frame, text=f"准确率: {accuracy:.2f}%", font=("微软雅黑", 14), bg="#f0f0f0").grid(row=3, column=0, sticky="w", pady=5)
        tk.Label(stats_frame, text=f"剩余错题: {len(self.progress['wrong_questions'])}", font=("微软雅黑", 14), bg="#f0f0f0").grid(row=4, column=0, sticky="w", pady=5)

        # 按钮框架
        btn_frame = tk.Frame(main_frame, bg="#f0f0f0")
        btn_frame.pack(pady=20)

        retry_btn = tk.Button(btn_frame, text="重新练习错题", command=self.retry_wrong_questions,
                              font=("微软雅黑", 12), bg="#FF9800", fg="white")
        retry_btn.pack(side=tk.LEFT, padx=10)

        new_btn = tk.Button(btn_frame, text="新练习", command=self.create_welcome_frame,
                            font=("微软雅黑", 12), bg="#4CAF50", fg="white")
        new_btn.pack(side=tk.LEFT, padx=10)

        exit_btn = tk.Button(btn_frame, text="退出", command=self.root.quit,
                             font=("微软雅黑", 12), bg="#F44336", fg="white")
        exit_btn.pack(side=tk.LEFT, padx=10)

    def retry_wrong_questions(self):
        """重新练习错题"""
        if not self.progress["wrong_questions"]:
            messagebox.showinfo("提示", "没有错题需要练习!")
            return

        self.question_order = self.progress["wrong_questions"][:]
        random.shuffle(self.question_order)
        self.current_index = 0
        self.show_question()

    def show_progress_management(self):
        """显示进度管理界面"""
        self.clear_frame()

        # 标题
        tk.Label(self.root, text="进度管理", font=("微软雅黑", 20, "bold"), bg="#f0f0f0").pack(pady=20)

        # 进度文件列表
        progress_files = []
        if os.path.exists(PROGRESS_DIR):
            progress_files = [f for f in os.listdir(PROGRESS_DIR) if f.endswith(".json")]

        if not progress_files:
            tk.Label(self.root, text="没有找到进度文件", font=("微软雅黑", 14), bg="#f0f0f0").pack(pady=20)
        else:
            list_frame = tk.Frame(self.root, bg="#f0f0f0")
            list_frame.pack(fill=tk.BOTH, expand=True, padx=50, pady=10)

            for i, file in enumerate(progress_files):
                file_frame = tk.Frame(list_frame, bg="#f0f0f0")
                file_frame.grid(row=i, column=0, sticky="ew", pady=5)

                tk.Label(file_frame, text=file, font=("微软雅黑", 11), bg="#f0f0f0").pack(side=tk.LEFT)

                del_btn = tk.Button(file_frame, text="删除", command=lambda f=file: self.delete_progress(f),
                                    font=("微软雅黑", 10), bg="#F44336", fg="white")
                del_btn.pack(side=tk.RIGHT, padx=10)

        # 按钮框架
        btn_frame = tk.Frame(self.root, bg="#f0f0f0")
        btn_frame.pack(pady=20)

        tk.Button(btn_frame, text="返回", command=self.create_welcome_frame,
                  font=("微软雅黑", 12), bg="#2196F3", fg="white").pack(side=tk.LEFT, padx=10)

        tk.Button(btn_frame, text="清空所有进度", command=self.clear_all_progress,
                  font=("微软雅黑", 12), bg="#F44336", fg="white").pack(side=tk.LEFT, padx=10)

    def delete_progress(self, file):
        """删除单个进度文件"""
        file_path = os.path.join(PROGRESS_DIR, file)
        try:
            os.remove(file_path)
            messagebox.showinfo("成功", "进度文件已删除")
            self.show_progress_management()  # 刷新列表
        except Exception as e:
            messagebox.showerror("错误", f"删除失败: {e}")

    def clear_all_progress(self):
        """清空所有进度"""
        if messagebox.askyesno("确认", "确定要清空所有进度吗？"):
            for file in os.listdir(PROGRESS_DIR):
                file_path = os.path.join(PROGRESS_DIR, file)
                try:
                    os.remove(file_path)
                except:
                    pass
            messagebox.showinfo("成功", "所有进度已清空")
            self.show_progress_management()  # 刷新列表

    def clear_frame(self):
        """清除当前框架内容"""
        for widget in self.root.winfo_children():
            widget.destroy()


if __name__ == "__main__":
    root = tk.Tk()
    app = ExamApp(root)
    root.mainloop()